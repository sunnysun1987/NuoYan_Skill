from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ivd_research.import_finding import import_finding


TITLE_KEYS = ["title", "题名", "标题", "文献标题"]
CONTENT_KEYS = ["abstract", "摘要", "summary", "内容", "备注", "人工摘录"]
URL_KEYS = ["url", "source_url", "链接", "来源链接", "详情页"]
DOI_KEYS = ["doi", "DOI"]
PMID_KEYS = ["pmid", "PMID"]
DATE_KEYS = ["publication_date", "发表日期", "日期", "年份"]


def _first(row: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _rows_from_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for cells in ws.iter_rows(min_row=2):
        row = {
            headers[index]: cell.value
            for index, cell in enumerate(cells)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def import_literature_table(task_dir: Path, path: Path, *, source: str = "csv_literature_import") -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _rows_from_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _rows_from_xlsx(path)
    else:
        raise ValueError("Only CSV/XLSX literature tables are supported.")
    imported = []
    for row in rows:
        title = _first(row, TITLE_KEYS)
        if not title:
            continue
        content = _first(row, CONTENT_KEYS) or title
        doi = _first(row, DOI_KEYS)
        pmid = _first(row, PMID_KEYS)
        identifier = pmid or doi
        result = import_finding(
            task_dir,
            title=title,
            source=source,
            source_url=_first(row, URL_KEYS),
            content=content,
            material_type="literature",
            identifier=identifier,
            publication_date=_first(row, DATE_KEYS),
            extra_raw_fields={
                "doi": doi,
                "pmid": pmid,
                "table_import_source": str(path),
                "table_row": row,
            },
        )
        imported.append(result)
    return {"imported_count": len(imported), "materials": imported}

