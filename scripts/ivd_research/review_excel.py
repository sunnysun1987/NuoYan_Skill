import json
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .jsonl import append_jsonl, read_jsonl, write_json
from .constants import EVIDENCE_STRENGTH_LABELS
from .quality import build_collection_alerts
from .project_profile import formal_scenarios_for
from .status import load_task


SHEET_BY_TYPE = {
    "regulatory": "监管资料",
    "competitor": "竞品注册",
    "standard": "现行标准",
    "patent": "专利",
    "literature": "文献",
    "local_import": "本地导入",
    "unknown": "待人工复核",
}

HEADERS = [
    "材料ID",
    "证据卡ID",
    "材料类型",
    "标题",
    "来源场景",
    "来源链接",
    "本地文件路径",
    "下载状态",
    "全文解析状态",
    "是否纳入报告",
    "一级标签",
    "二级标签",
    "证据强度",
    "可信度说明",
    "摘要修订",
    "报告用途修订",
    "人工备注",
    "复核状态",
]

LITERATURE_HEADERS = [
    "材料ID",
    "证据卡ID",
    "来源场景",
    "题名",
    "PMID",
    "PMCID",
    "DOI",
    "期刊",
    "发表日期",
    "摘要状态",
    "结构化Abstract",
    "全文状态",
    "PDF状态",
    "来源链接",
    "本地文件路径",
    "补证建议",
    "复核状态",
]

TAXONOMY = [
    "临床意义",
    "临床应用定位",
    "目标使用场景",
    "目标人群",
    "诊疗路径",
    "金标准与参照方法",
    "指南与共识",
    "专家和组织意见",
    "市场准入与收费",
    "市场定位",
    "竞争格局",
    "出口与注册",
    "技术可行性",
    "参考物质",
    "安全要求",
    "原材料可获得性",
]


def first_local_path(material: dict[str, Any]) -> str:
    if material.get("download_files"):
        return material["download_files"][0].get("relative_path", "")
    return material.get("collection_path", {}).get("stored_file", "")


def evidence_by_material(task_dir: Path) -> dict[str, dict[str, Any]]:
    evidence = {}
    for card in read_jsonl(task_dir / "data" / "evidence_cards.jsonl"):
        evidence.setdefault(card.get("material_id", ""), card)
    return evidence


def material_ids(task_dir: Path) -> set[str]:
    return {
        material["material_id"]
        for material in read_jsonl(task_dir / "data" / "materials.jsonl")
        if material.get("material_id")
    }


def split_labels(value: Any) -> list[str]:
    if value is None:
        return []
    labels = []
    for part in str(value).replace(";", "；").split("；"):
        label = part.strip()
        if label:
            labels.append(label)
    return labels


def parse_include(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    return normalized in {"是", "yes", "true", "1", "y"}


def normalize_evidence_strength(value: Any) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return "needs_review"
    if normalized in EVIDENCE_STRENGTH_LABELS:
        return normalized
    for key, label in EVIDENCE_STRENGTH_LABELS.items():
        if normalized == label:
            return key
    return normalized


def reviewed_card_from_values(card: dict[str, Any], values: dict[str, Any]) -> dict[str, Any]:
    reviewed = dict(card)
    reviewed["include_in_report"] = parse_include(values.get("是否纳入报告"))
    reviewed["taxonomy_tags"] = split_labels(values.get("一级标签")) + split_labels(
        values.get("二级标签")
    )
    reviewed["evidence_strength"] = normalize_evidence_strength(values.get("证据强度"))
    summary = values.get("摘要修订")
    if summary is not None:
        reviewed["summary"] = str(summary)
    report_usage = values.get("报告用途修订")
    if report_usage is not None:
        reviewed["report_usage"] = str(report_usage)
    reviewed["manual_review"] = {
        "manual_notes": str(values.get("人工备注") or ""),
        "review_status": str(values.get("复核状态") or ""),
    }
    review_status = reviewed["manual_review"]["review_status"].strip().lower()
    if review_status in {"已复核", "reviewed", "approved"}:
        reviewed["needs_review"] = False
        reviewed["review_reasons"] = []
    return reviewed


def write_reviewed_cards(task_dir: Path, reviewed_cards: list[dict[str, Any]]) -> None:
    path = task_dir / "data" / "reviewed_evidence_cards.jsonl"
    by_id = {
        card["evidence_card_id"]: card
        for card in read_jsonl(path)
        if card.get("evidence_card_id")
    }
    for card in reviewed_cards:
        card_id = card.get("evidence_card_id")
        if card_id:
            by_id[card_id] = card

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for card in by_id.values():
            handle.write(json.dumps(card, ensure_ascii=False, sort_keys=True) + "\n")


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    for index, column in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 12), 38)


def export_review(task_dir: Path) -> dict:
    materials = list(read_jsonl(task_dir / "data" / "materials.jsonl"))
    evidence = evidence_by_material(task_dir)
    task = load_task(task_dir)
    scenario_statuses = [
        status.model_dump(mode="json")
        for status in (task.scenario_statuses or {}).values()
    ]
    required_scenarios = formal_scenarios_for(task.model_dump(mode="json"))
    collection_alerts = build_collection_alerts(
        materials=materials,
        evidence_cards=list(evidence.values()),
        scenario_statuses=scenario_statuses,
        required_scenario_ids=required_scenarios,
    )
    wb = Workbook()
    wb.remove(wb.active)

    alerts_ws = wb.create_sheet("采集异常")
    alerts_ws.append(["项目", "值"])
    alerts_ws.append(["告警级别", collection_alerts["level"]])
    alerts_ws.append(["告警标题", collection_alerts["headline"]])
    alerts_ws.append(["材料数", collection_alerts["material_count"]])
    alerts_ws.append(["证据卡数", collection_alerts["evidence_card_count"]])
    alerts_ws.append(["失败场景数", collection_alerts["failed_count"]])
    alerts_ws.append(["未启动场景数", collection_alerts["not_started_count"]])
    alerts_ws.append([])
    alerts_ws.append(["必须先处理的问题", ""])
    for message in collection_alerts["critical_messages"]:
        alerts_ws.append(["问题", message])
    alerts_ws.append([])
    alerts_ws.append(["待补证或待确认事项", ""])
    for message in collection_alerts["warning_messages"]:
        alerts_ws.append(["事项", message])

    for sheet in list(dict.fromkeys(SHEET_BY_TYPE.values())):
        ws = wb.create_sheet(sheet)
        ws.append(HEADERS)

    dictionary = wb.create_sheet("标签字典")
    dictionary.append(["一级标签", "二级标签"])
    for primary in TAXONOMY:
        dictionary.append([primary, ""])

    instructions = wb.create_sheet("填写说明")
    instructions.append(["说明"])
    instructions.append(["请不要修改材料ID和证据卡ID。多标签请用中文分号；分隔。"])
    instructions.append(["是否纳入报告建议填写：是 / 否。复核状态建议填写：未复核 / 已复核 / 需补充。"])

    literature_review = wb.create_sheet("文献检索")
    literature_review.append(LITERATURE_HEADERS)
    metric_ws = wb.create_sheet("指标事实")
    metric_ws.append(
        [
            "指标事实ID",
            "指标类型",
            "数值",
            "材料ID",
            "证据卡ID",
            "样本类型",
            "平台",
            "参照方法",
            "队列/人群",
            "比较对象",
            "原文线索",
            "来源位置",
            "复核状态",
        ]
    )

    for material in materials:
        material_type = material.get("material_type", "unknown")
        sheet = SHEET_BY_TYPE.get(material_type, SHEET_BY_TYPE["unknown"])
        card = evidence.get(material.get("material_id", ""), {})
        ws = wb[sheet]
        ws.append(
            [
                material.get("material_id", ""),
                card.get("evidence_card_id", ""),
                material_type,
                material.get("title", ""),
                material.get("source_scenario", ""),
                material.get("source_url", ""),
                first_local_path(material),
                material.get("download_status", ""),
                material.get("extracted_text_status", ""),
                "是" if card.get("include_in_report") else "否",
                "；".join(card.get("taxonomy_tags", [])[:1]),
                "；".join(card.get("taxonomy_tags", [])[1:]),
                card.get("evidence_strength", "待人工复核"),
                card.get("confidence_level", ""),
                card.get("summary", ""),
                card.get("report_usage", ""),
                "",
                "未复核",
            ]
        )
        if material_type == "literature":
            raw = material.get("raw_fields") or {}
            abstract_status = "有摘要" if raw.get("abstract") else "无摘要/未解析"
            structured_abstract = _structured_abstract_text(raw)
            fulltext_status = raw.get("fulltext_status") or material.get("extracted_text_status", "")
            pdf_status = raw.get("pdf_status") or material.get("download_status", "")
            literature_review.append(
                [
                    material.get("material_id", ""),
                    card.get("evidence_card_id", ""),
                    material.get("source_scenario", ""),
                    material.get("title", ""),
                    raw.get("pmid", ""),
                    raw.get("pmcid", ""),
                    raw.get("doi", ""),
                    raw.get("journal", "") or raw.get("journal_iso", ""),
                    raw.get("publication_date", ""),
                    abstract_status,
                    structured_abstract,
                    fulltext_status,
                    pdf_status,
                    material.get("source_url", ""),
                    first_local_path(material),
                    _literature_gap_action(raw, material),
                    "未复核",
                ]
            )

    for fact in read_jsonl(task_dir / "knowledge" / "metric_facts.jsonl"):
        metric_ws.append(
            [
                fact.get("metric_fact_id", ""),
                fact.get("metric_type", ""),
                fact.get("value", ""),
                fact.get("material_id", ""),
                fact.get("evidence_card_id", ""),
                fact.get("sample_type", ""),
                fact.get("platform", ""),
                fact.get("reference_standard", ""),
                fact.get("cohort", ""),
                fact.get("comparator", ""),
                fact.get("excerpt", ""),
                fact.get("source_location", ""),
                "未复核",
            ]
        )

    for ws in wb.worksheets:
        style_sheet(ws)

    path = task_dir / "review" / "evidence_review_v001.xlsx"
    standard_path = task_dir / "review" / "01_证据审阅与补证任务表.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    shutil.copy2(path, standard_path)
    return {
        "review_path": str(standard_path),
        "legacy_review_path": str(path),
        "material_count": len(materials),
    }


def _literature_gap_action(raw: dict[str, Any], material: dict[str, Any]) -> str:
    actions = []
    if not raw.get("pmid") and material.get("source_scenario") == "pubmed_literature":
        actions.append("补查 PMID")
    if not raw.get("doi"):
        actions.append("补查 DOI")
    if not raw.get("abstract"):
        actions.append("补充摘要")
    if raw.get("pmcid") and raw.get("fulltext_status") not in {"completed"}:
        actions.append("补查 PMC 全文")
    if raw.get("pmcid") and raw.get("pdf_status") not in {"downloaded", "not_available"}:
        actions.append("复核 PMC PDF 下载")
    if not raw.get("pmcid") and material.get("source_scenario") == "pubmed_literature":
        actions.append("判断是否存在开放全文")
    return "；".join(actions) if actions else "无明确补证动作，待人工确认采用价值"


def _structured_abstract_text(raw: dict[str, Any]) -> str:
    sections = raw.get("abstract_sections") or []
    lines: list[str] = []
    if isinstance(sections, list):
        for section in sections:
            if not isinstance(section, dict):
                continue
            label = str(section.get("label") or "").strip() or "Abstract"
            text = " ".join(str(section.get("text") or "").split())
            if text:
                lines.append(f"{label}: {text}")
    if not lines:
        abstract = " ".join(str(raw.get("abstract") or "").split())
        if abstract:
            lines.append(abstract)
    keywords = raw.get("keywords") or []
    if isinstance(keywords, list):
        keyword_text = "；".join(str(item).strip() for item in keywords if str(item).strip())
    else:
        keyword_text = str(keywords or "").strip()
    if keyword_text:
        lines.append(f"Keywords: {keyword_text}")
    return "\n".join(lines)


def import_review(task_dir: Path, workbook_path: Path) -> dict:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Review workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path)
    errors = []
    imports = []
    reviewed_cards = []
    known_material_ids = material_ids(task_dir)
    evidence = evidence_by_material(task_dir)
    taxonomy = set(TAXONOMY)
    for sheet_name in SHEET_BY_TYPE.values():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = dict(zip(headers, row))
            if not any(values.values()):
                continue
            material_id = str(values.get("材料ID") or "").strip()
            if not material_id:
                errors.append(
                    {"sheet": sheet_name, "row": row_index, "error": "缺少材料ID"}
                )
                continue
            if material_id not in known_material_ids:
                errors.append(
                    {
                        "sheet": sheet_name,
                        "row": row_index,
                        "error": f"未知材料ID：{material_id}",
                    }
                )
            labels = split_labels(values.get("一级标签")) + split_labels(values.get("二级标签"))
            invalid_labels = [label for label in labels if label not in taxonomy]
            if invalid_labels:
                errors.append(
                    {
                        "sheet": sheet_name,
                        "row": row_index,
                        "error": f"非法标签：{'；'.join(invalid_labels)}",
                    }
                )

            import_row = {
                "sheet": sheet_name,
                "row": row_index,
                "material_id": material_id,
                "evidence_card_id": values.get("证据卡ID") or "",
                "values": values,
            }
            imports.append(import_row)
            if material_id in evidence:
                reviewed_cards.append(reviewed_card_from_values(evidence[material_id], values))
            elif material_id in known_material_ids:
                errors.append(
                    {
                        "sheet": sheet_name,
                        "row": row_index,
                        "error": f"Missing source evidence card for material_id: {material_id}",
                    }
                )

    if errors:
        report = {"ok": False, "imported_count": 0, "errors": errors}
        write_json(task_dir / "review" / "import_validation_v001.json", report)
        return report

    for row in imports:
        append_jsonl(task_dir / "data" / "review_imports.jsonl", row)
    write_reviewed_cards(task_dir, reviewed_cards)

    report = {"ok": True, "imported_count": len(imports), "errors": errors}
    write_json(task_dir / "review" / "import_validation_v001.json", report)
    return report
