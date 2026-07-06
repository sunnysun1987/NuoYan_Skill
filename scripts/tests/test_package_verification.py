import json
from pathlib import Path

import pytest
from bs4 import BeautifulSoup

from ivd_research.evidence import (
    build_draft_evidence_card,
    export_evidence_card_files,
    generate_draft_evidence_cards,
)
from ivd_research.jsonl import append_jsonl, write_json
from ivd_research.models import Material
from ivd_research.package import (
    FORMAL_SCENARIOS,
    build_standard_delivery,
    requires_life_science_research,
    verify_package,
)
from ivd_research.review_excel import export_review, import_review
from ivd_research.source_adapters.life_science_research_bridge import import_life_science_findings
from ivd_research.status import init_task
from ivd_research.confirmations import update_confirmations


FULL_CONFIRMATIONS = {
    "task_info": True,
    "keyword_pool": True,
    "collection_scope": True,
    "primary_query": "血浆 p-tau217 阿尔茨海默病 体外诊断",
    "english_keywords": "plasma p-tau217 Alzheimer disease IVD",
    "sample_type": "血浆",
    "platform": "化学发光",
    "methodology": True,
    "intended_use": "阿尔茨海默病辅助诊断",
    "target_region": "中国",
    "competitor_scope": "NMPA 已注册同类产品",
    "patent_scope": True,
}


def _task_dir(tmp_path: Path) -> Path:
    state = init_task("p-tau217 稳定性测试", tmp_path)
    return Path(state.task_dir)


def _write_single_material_and_card(task_dir: Path) -> None:
    text_path = task_dir / "extracted_text" / "literature" / "MAT-000001.txt"
    text_path.parent.mkdir(parents=True, exist_ok=True)
    text_path.write_text(
        "PMID：12345678\nDOI：10.1000/test\n摘要：p-tau217 与 AD 病理相关。",
        encoding="utf-8",
    )
    material = Material(
        material_id="MAT-000001",
        task_id="TEST",
        source_scenario="pubmed_literature",
        material_type="literature",
        title="Plasma p-tau217 for Alzheimer disease diagnosis",
        source_url="https://pubmed.ncbi.nlm.nih.gov/12345678/",
        search_keyword_or_query="plasma p-tau217 Alzheimer disease",
        collection_path={"scenario_id": "pubmed_literature"},
        collection_time="2026-06-16T00:00:00+08:00",
        adapter_id="pubmed_literature",
        adapter_version="2.0.0",
        raw_fields={
            "pmid": "12345678",
            "doi": "10.1000/test",
            "journal": "Journal of Test Medicine",
            "publication_date": "2026-06-16",
            "abstract": "p-tau217 is associated with Alzheimer pathology.",
            "fulltext_status": "completed",
            "pdf_status": "not_available",
        },
        extracted_text_status="completed",
        extracted_text_path=str(text_path.relative_to(task_dir)),
    )
    append_jsonl(task_dir / "data" / "materials.jsonl", material.model_dump(mode="json"))
    card = build_draft_evidence_card(
        task_dir,
        material.model_dump(mode="json"),
        "EC-000001",
    )
    card_payload = card.model_dump(mode="json")
    append_jsonl(task_dir / "data" / "evidence_cards.jsonl", card_payload)
    export_evidence_card_files(task_dir, card_payload)


def _mark_formal_scenarios(task_dir: Path, status: str = "no_results") -> None:
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    for scenario_id in FORMAL_SCENARIOS:
        task["scenario_statuses"][scenario_id]["status"] = status
        task["scenario_statuses"][scenario_id]["last_message"] = "离线测试：已记录明确状态。"
    task["scenario_statuses"]["pubmed_literature"]["status"] = "completed"
    task["scenario_statuses"]["pubmed_literature"]["material_count"] = 1
    write_json(task_dir / "task.json", task)


def _import_complete_life_science_coverage(task_dir: Path) -> None:
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    findings = []
    seed_rows = [
        ("UniProt", "protein", "MAPT"),
        ("Open Targets", "target", "MAPT"),
        ("ClinicalTrials.gov", "clinical", "p-tau217"),
        ("ClinicalTrials.gov", "clinical", "p-tau181"),
        ("STRING", "network", "MAPT"),
        ("Reactome", "pathway", "tau protein binding"),
        ("QuickGO", "pathway", "microtubule binding"),
        ("Human Protein Atlas", "protein", "MAPT brain expression"),
        ("GWAS Catalog", "genetics", "Alzheimer disease"),
        ("NCBI Gene", "target", "MAPT"),
        ("EFO/OLS", "disease", "Alzheimer disease"),
        ("ClinicalTrials.gov", "clinical", "plasma biomarker"),
    ]
    for index, (database, lane, entity) in enumerate(seed_rows, start=1):
        findings.append(
            {
                "source_database": database,
                "evidence_lane": lane,
                "entity": entity,
                "query": "plasma p-tau217 Alzheimer disease",
                "result_summary": f"{database} evidence row {index} for {entity}.",
                "source_url": f"https://example.org/life-science/{index}",
                "identifier": f"LS-{index:03d}",
            }
        )
    import_life_science_findings(
        task["task_id"],
        task_dir,
        findings,
        query="plasma p-tau217 Alzheimer disease",
    )


def test_verify_package_keeps_incomplete_scope_as_not_business_ready(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    _write_single_material_and_card(task_dir)
    export_review(task_dir)
    build_standard_delivery(task_dir)

    result = verify_package(task_dir)

    assert result["delivery_artifacts_ready"] is True
    assert (task_dir / "交付目录" / "02_证据卡" / "EC-000001.md").exists()
    assert result["business_ready"] is False
    assert "task_info" in result["missing_confirmations"]
    assert result["final_review_ready"] is False


def test_life_science_requirement_detects_ad_without_lead_false_positive():
    assert requires_life_science_research(
        {
            "topic": "AD p-tau181 血液标志物",
            "confirmations": {"english_keywords": "AD plasma biomarker"},
        }
    )
    assert not requires_life_science_research(
        {
            "topic": "lead time workflow improvement",
            "confirmations": {"english_keywords": "lead time process validation"},
        }
    )


def test_standard_delivery_report_has_drilldown_navigation_and_metric_definitions(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    _write_single_material_and_card(task_dir)
    export_review(task_dir)
    build_standard_delivery(task_dir)

    html = (task_dir / "交付目录" / "00_立项调研综合报告.html").read_text(encoding="utf-8")

    assert "章节目录" in html
    assert "analysis-nav-card" in html
    assert "点击后跳转到对应章节" in html
    assert 'data-jump-tab="tab-screening"' in html
    assert 'data-jump-target="screening-card-list"' in html
    assert 'data-tab-target="tab-metrics">指标事实' in html
    assert 'data-jump-tab="tab-metrics" data-jump-target="metric-facts"' in html
    assert 'data-jump-target="metric-facts"' in html
    assert 'id="tab-metrics"' in html
    assert "已写入材料库的原始资料总数" in html
    assert "从证据中结构化抽取的样本量" in html
    assert "口径：" not in html
    assert "中文阅读覆盖" not in html
    assert "metric-global-search" in html
    assert 'data-metric-filter="metric"' in html
    assert "指标字段" in html
    assert "数值字段" in html
    assert "analysis-evidence-table" in html
    assert "analysis-evidence-panel" in html
    assert 'data-page-size="8"' in html
    assert "依据清单" in html
    assert "相关内容原文" in html
    assert "先看结论" in html
    assert "研发定位" in html
    assert html.index('id="tab-metrics"') < html.index('id="tab-core"')

    soup = BeautifulSoup(html, "html.parser")
    reading_panel = soup.select_one("#tab-reading")
    metrics_panel = soup.select_one("#tab-metrics")
    assert reading_panel is not None
    assert metrics_panel is not None
    assert reading_panel.select_one("#metric-facts") is None
    assert metrics_panel.select_one("#metric-facts") is not None


def test_verify_package_requires_fallback_for_failed_formal_scenarios(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    export_review(task_dir)
    build_standard_delivery(task_dir)
    _mark_formal_scenarios(task_dir)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    task["scenario_statuses"]["patenthub_patents"]["status"] = "permission_required"
    task["scenario_statuses"]["patenthub_patents"]["last_message"] = "PatentHub 需要登录。"
    write_json(task_dir / "task.json", task)

    result = verify_package(task_dir)

    assert result["fallback_ready"] is False
    assert result["scenario_coverage_ready"] is False
    assert result["business_ready"] is False
    assert any("PatentHub" in warning for warning in result["warnings"])


def test_verify_package_accepts_reviewed_complete_offline_package(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    _import_complete_life_science_coverage(task_dir)
    generate_draft_evidence_cards(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    row = 2
    for header, value in {
        "是否纳入报告": "是",
        "一级标签": "临床意义",
        "证据强度": "moderate",
        "复核状态": "已复核",
    }.items():
        ws.cell(row=row, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    import_result = import_review(task_dir, workbook)
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert import_result["ok"] is True
    assert result["delivery_artifacts_ready"] is True
    assert result["final_review_ready"] is True
    assert result["search_profile_ready"] is True
    assert result["scenario_coverage_ready"] is True
    assert result["fallback_ready"] is True
    assert result["network_ready"] is True
    assert result["business_ready"] is True


def test_verify_package_requires_life_science_for_biomarker_projects(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    for header, value in {
        "是否纳入报告": "是",
        "一级标签": "临床意义",
        "证据强度": "moderate",
        "复核状态": "已复核",
    }.items():
        ws.cell(row=2, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    assert import_review(task_dir, workbook)["ok"] is True
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert result["life_science_coverage"]["required"] is True
    assert result["scenario_coverage_ready"] is False
    assert result["business_ready"] is False
    assert any("尚未导入 life-science-research" in warning for warning in result["warnings"])


def test_verify_package_blocks_shallow_life_science_coverage(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    import_life_science_findings(
        json.loads((task_dir / "task.json").read_text(encoding="utf-8"))["task_id"],
        task_dir,
        [
            {
                "source_database": "UniProt",
                "evidence_lane": "protein",
                "entity": "MAPT",
                "result_summary": "Single protein finding.",
            }
        ],
        query="MAPT Alzheimer disease",
    )
    generate_draft_evidence_cards(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    for row in range(2, ws.max_row + 1):
        for header, value in {
            "是否纳入报告": "是",
            "一级标签": "临床意义",
            "证据强度": "moderate",
            "复核状态": "已复核",
        }.items():
            ws.cell(row=row, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    assert import_review(task_dir, workbook)["ok"] is True
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert result["life_science_coverage"]["material_count"] == 1
    assert result["scenario_coverage_ready"] is False
    assert any("插件证据数量不足" in warning for warning in result["warnings"])
    assert any("插件数据库覆盖不足" in warning for warning in result["warnings"])
    assert any("插件证据通道覆盖不足" in warning for warning in result["warnings"])


def test_verify_package_blocks_unresolved_network_after_failed_preflight(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    task["scenario_statuses"]["pubmed_literature"]["status"] = "not_started"
    task["scenario_statuses"]["pubmed_literature"]["material_count"] = 0
    task["scenario_statuses"]["pubmed_literature"]["last_message"] = ""
    write_json(task_dir / "task.json", task)
    append_jsonl(
        task_dir / "logs" / "events.jsonl",
        {
            "event": "network_preflight",
            "network_ok": False,
            "probes": [
                {
                    "id": "pubmed",
                    "label_zh": "PubMed",
                    "python_dns_error": "No DNS configuration available",
                }
            ],
        },
    )

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    for header, value in {
        "是否纳入报告": "是",
        "一级标签": "临床意义",
        "证据强度": "moderate",
        "复核状态": "已复核",
    }.items():
        ws.cell(row=2, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    assert import_review(task_dir, workbook)["ok"] is True
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert result["network_ready"] is False
    assert result["business_ready"] is False
    assert result["network_unresolved_scenarios"] == [
        {
            "scenario_id": "pubmed_literature",
            "status": "not_started",
            "material_count": 0,
            "last_message": "",
        }
    ]


def test_update_confirmations_rejects_unknown_keys(tmp_path: Path):
    task_dir = _task_dir(tmp_path)

    with pytest.raises(ValueError, match="Unknown confirmation key: target_market"):
        update_confirmations(task_dir, {"target_market": "中国"})
